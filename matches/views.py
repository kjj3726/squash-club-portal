import random
import csv
from django.http import HttpResponse
from django.http import JsonResponse
from datetime import date, timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
# 🌟 여기서 F를 정상적으로 불러옵니다.
from django.db.models import Q, F, Count, Window
# 🌟 [수정] 순차 번호를 매기기 위한 RowNumber 함수를 불러옵니다.
from django.db.models.functions import RowNumber
from .models import Profile, MonthlyMeet, Match, Notice, NoticeComment
from django.contrib.auth.models import User
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth import update_session_auth_hash 
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from django.utils.html import escape


# --- (신규) 권한 체크용 헬퍼 함수 ---
def is_manager(user):
    """총관리자(is_superuser)이거나 사장님(is_owner)인지 확인"""
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    if hasattr(user, 'profile') and user.profile.is_owner:
        return True
    return False

# 1. 핸디캡 계산 엔진 (규칙 업데이트 및 반환 로직 완료)
def calculate_handicap_logic(p1, p2):
    # 1. 조별 기본 실력 세팅 (A조=6, B조=3, C조=0)
    tier_power = {'A': 6, 'B': 3, 'C': 0}
    p1_power = tier_power.get(p1.group, 0)
    p2_power = tier_power.get(p2.group, 0)

    # 2. 성별 실력 세팅 (남성이 2점 더 강함 -> 결과적으로 여성이 2점 이득을 봄)
    if p1.gender == 'M':
        p1_power += 2
    if p2.gender == 'M':
        p2_power += 2

    # 3. 강자/약자 판별 및 순수 핸디캡 점수 도출
    if p1_power > p2_power:
        strong_p, weak_p = p1, p2
        final_handicap = p1_power - p2_power
    elif p2_power > p1_power:
        strong_p, weak_p = p2, p1
        final_handicap = p2_power - p1_power
    else:
        return 0 # 전력이 완벽하게 동일함 (핸디캡 0)

    # 4. 전적 차감 (진짜 약자가 강자를 이긴 횟수만큼 차감)
    win_count = Match.objects.filter(
        (Q(player1=weak_p, player2=strong_p, p1_score__gt=F('p2_score')) |
         Q(player1=strong_p, player2=weak_p, p2_score__gt=F('p1_score'))),
        is_completed=True
    ).count()

    # 5. 최종 핸디캡 확정 (승리 횟수만큼 차감하되 최소 0점 보장)
    final_handicap = max(0, final_handicap - win_count)
    
    return final_handicap

# 2. 그룹별 승률 계산 헬퍼 함수
def get_top_players(group_name):
    profiles = Profile.objects.filter(group=group_name, is_owner=False, is_guest=False)
    stats = []
    for p in profiles:
        matches = Match.objects.filter(
            Q(player1=p) | Q(player2=p), 
            is_completed=True,
            meet__is_finalized=True 
        )
        total = matches.count()
        wins = 0
        for m in matches:
            if m.player1 == p and (m.p1_score or 0) > (m.p2_score or 0): wins += 1
            if m.player2 == p and (m.p2_score or 0) > (m.p1_score or 0): wins += 1
            
        win_rate = int((wins / total * 100)) if total > 0 else 0
        losses = total - wins  # 🌟 패배 횟수 계산
        
        if total > 0:
            # 🌟 wins(승)와 losses(패) 데이터를 화면으로 넘겨줍니다.
            stats.append({'profile': p, 'win_rate': win_rate, 'total': total, 'wins': wins, 'losses': losses})
            
    stats.sort(key=lambda x: x['win_rate'], reverse=True)
    return stats[:3] # 🌟 1위부터 3위까지만 자릅니다

# 3. 메인 포털 대시보드 뷰
@login_required(login_url='login')
def dashboard(request):
    # 🌟 [수정됨] 당일 자정(23:59)까지는 마감 여부와 상관없이 오늘 모임을 메인에 노출합니다.
    meet = MonthlyMeet.objects.filter(date=date.today()).first()
    
    # 만약 오늘 진행된 모임이 없다면, 기존처럼 마감되지 않은 가장 최근 예정 모임을 가져옵니다.
    if not meet:
        meet = MonthlyMeet.objects.filter(is_finalized=False).order_by('-date').first()
    
    # 용도에 맞춰 회원 명단을 분리 (사장님 완벽 제외)
    all_profiles = Profile.objects.filter(is_owner=False).order_by('name')
    regular_profiles = Profile.objects.filter(is_owner=False, is_guest=False).order_by('name')
    guest_profiles = Profile.objects.filter(is_owner=False, is_guest=True).order_by('name')
    ranking_profiles = Profile.objects.filter(is_owner=False, is_guest=False)

    # 🌟 [수정] DB에서 목록을 가져온 후, 파이썬에서 순차 번호를 부여합니다.
    notices_qs = Notice.objects.annotate(
        comment_count=Count('comments')
    ).order_by('-is_important', '-created_at')[:5]
    
    # 🌟 일반 공지사항의 총 개수만 계산합니다.
    total_normal_notices = Notice.objects.filter(is_important=False).count()
    notices = []
    normal_index = 0
    for notice in notices_qs:
        if notice.is_important:
            notice.display_id = "공지"
        else:
            notice.display_id = total_normal_notices - normal_index
            normal_index += 1
        notices.append(notice)
    top_a = get_top_players('A')
    top_b = get_top_players('B')
    top_c = get_top_players('C')

    all_stats_sorted = []
    for p in ranking_profiles:
        p_matches = Match.objects.filter(Q(player1=p) | Q(player2=p), is_completed=True)
        p_total = p_matches.count()
        
        p_wins = sum(1 for m in p_matches if (m.player1 == p and (m.p1_score or 0) > (m.p2_score or 0)) or (m.player2 == p and (m.p2_score or 0) > (m.p1_score or 0)))
        p_losses = p_total - p_wins
        p_win_rate = int((p_wins / p_total * 100)) if p_total > 0 else 0
        
        all_stats_sorted.append({
            'profile': p,
            'win_rate': p_win_rate,
            'wins': p_wins,
            'losses': p_losses
        })
        
    all_stats_sorted.sort(key=lambda x: x['win_rate'], reverse=True)
    
    all_a = [s for s in all_stats_sorted if s['profile'].group == 'A']
    all_b = [s for s in all_stats_sorted if s['profile'].group == 'B']
    all_c = [s for s in all_stats_sorted if s['profile'].group == 'C']

    user_stat = None
    if request.user.is_authenticated and hasattr(request.user, 'profile'):
        p = request.user.profile
        matches = Match.objects.filter(Q(player1=p) | Q(player2=p), is_completed=True)
        total = matches.count()
        wins = sum(1 for m in matches if (m.player1 == p and (m.p1_score or 0) > (m.p2_score or 0)) or (m.player2 == p and (m.p2_score or 0) > (m.p1_score or 0)))
        win_rate = int((wins / total * 100)) if total > 0 else 0
        user_stat = {'win_rate': win_rate, 'total': total, 'wins': wins}

    login_failed_username = request.session.pop('login_failed_username', '')

    member_data_all = []
    member_data_regular = []
    member_data_guest = []
    
    for p in all_profiles:
        participation_count = Match.objects.filter(
            Q(player1=p) | Q(player2=p), 
            meet__is_finalized=True
        ).values('meet').distinct().count()
        
        data_node = {'profile': p, 'attendance_count': participation_count}
        member_data_all.append(data_node)
        if p.is_guest:
            member_data_guest.append(data_node)
        else:
            member_data_regular.append(data_node)
            
    member_data_all.sort(key=lambda x: x['attendance_count'], reverse=True)
    member_data_regular.sort(key=lambda x: x['attendance_count'], reverse=True)
    member_data_guest.sort(key=lambda x: x['attendance_count'], reverse=True)

    # 🌟 월별 모임 히스토리 데이터 구성
    if is_manager(request.user):
        # 관리자/사장님은 모든 모임 기록을 확인 가능
        all_meets = MonthlyMeet.objects.all().order_by('-date')
    else:
        # 일반 유저는 오늘(당일) 진행된 모임만 확인 가능
        all_meets = MonthlyMeet.objects.filter(date=date.today()).order_by('-date')
        
    history_data = []
    for m in all_meets:
        m_matches = Match.objects.filter(meet=m).order_by('court', 'id')
        history_data.append({
            'meet': m,
            'matches': m_matches,
            'is_latest': (m == meet) # 현재 메인에 떠 있는 최신 모임인지 구분
        })

    # 🌟 [신규 추가] 경기 누적 상태 데이터 (권한별 분리)
    if is_manager(request.user):
        manager_history_regular = []
        manager_history_guest = []
        for p in all_profiles:
            p_matches = Match.objects.filter(
                Q(player1=p) | Q(player2=p), 
                is_completed=True
            ).select_related('meet', 'player1', 'player2').order_by('-meet__date', '-id')
            
            if p_matches.exists():
                total = p_matches.count()
                
                # 🌟 승패 계산 로직 추가
                wins = 0
                for m in p_matches:
                    if m.player1 == p and (m.p1_score or 0) > (m.p2_score or 0):
                        wins += 1
                    elif m.player2 == p and (m.p2_score or 0) > (m.p1_score or 0):
                        wins += 1
                losses = total - wins

                data_node = {
                    'profile': p,
                    'matches': p_matches,
                    'total_count': total,
                    'wins': wins,        # 🌟 계산된 승리 전달
                    'losses': losses     # 🌟 계산된 패배 전달
                }
                
                if p.is_guest:
                    manager_history_guest.append(data_node)
                else:
                    manager_history_regular.append(data_node)
        
        manager_history_regular.sort(key=lambda x: x['total_count'], reverse=True)
        manager_history_guest.sort(key=lambda x: x['total_count'], reverse=True)
        user_history = None
        
    else:
        # 일반 유저: 자신이 속한 완료된 경기만 가져오기 (기존과 동일)
        manager_history_regular = None
        manager_history_guest = None
        user_history = None
        if hasattr(request.user, 'profile'):
            user_p = request.user.profile
            user_history = Match.objects.filter(
                Q(player1=user_p) | Q(player2=user_p), 
                is_completed=True
            ).select_related('meet', 'player1', 'player2').order_by('-meet__date', '-id')

    context = {
        'meet': meet,
        'all_profiles': all_profiles,
        'regular_profiles': regular_profiles,
        'guest_profiles': guest_profiles,
        'all_stats_sorted': all_stats_sorted, 
        'all_a': all_a, 
        'all_b': all_b, 
        'all_c': all_c, 
        'notices': notices,
        'top_a': top_a, 
        'top_b': top_b, 
        'top_c': top_c,
        'user_stat': user_stat,
        'is_manager': is_manager(request.user),
        'login_failed_username': login_failed_username,
        'member_data_all': member_data_all,
        'member_data_regular': member_data_regular,
        'member_data_guest': member_data_guest,
        'attendance_rank': member_data_regular[:5], 
        'history_data': history_data, 
        # 🌟 누적 경기 데이터 전달
        'manager_history_regular': manager_history_regular,
        'manager_history_guest': manager_history_guest,
        'user_history': user_history,
    }

    if meet:
        # 🌟 1. 매치 목록을 한 번에 가져옵니다.
        all_matches = Match.objects.filter(meet=meet).order_by('court', 'id')
        
        matches_2 = []
        matches_3 = []
        completed = []
        
        # 🌟 2. HTML이 헷갈리지 않도록 '진짜 약자(수혜자)'를 계산해서 붙여주고 리스트를 분배합니다.
        for m in all_matches:
            tier_power = {'A': 6, 'B': 3, 'C': 0}
            p1_power = tier_power.get(m.player1.group, 0) + (2 if m.player1.gender == 'M' else 0)
            p2_power = tier_power.get(m.player2.group, 0) + (2 if m.player2.gender == 'M' else 0)
            
            # 전력이 낮은 쪽(약자)을 receiver(수혜자)로 지정
            if p1_power > p2_power:
                m.receiver = m.player2
            elif p2_power > p1_power:
                m.receiver = m.player1
            else:
                m.receiver = None

            # 상태 및 코트별로 리스트 분배
            if m.is_completed:
                completed.append(m)
            else:
                if m.court == 2:
                    matches_2.append(m)
                elif m.court == 3:
                    matches_3.append(m)

        context['matches_2'] = matches_2
        context['matches_3'] = matches_3
        context['completed'] = completed
        context['all_matches'] = all_matches
        
        # 🌟 신규: 오늘 대진표에 있는 player1, player2의 ID를 모두 모아서 중복을 제거한 뒤, 그 사람들만 전달
        p1_ids = all_matches.values_list('player1_id', flat=True)
        p2_ids = all_matches.values_list('player2_id', flat=True)
        participating_ids = set(list(p1_ids) + list(p2_ids))
        context['participating_profiles'] = Profile.objects.filter(id__in=participating_ids).order_by('name')

    return render(request, 'matches/dashboard.html', context)

# 4. 팝업창 연동: 모임 생성 및 랜덤 배치 (게스트 처리 및 날짜 중복 방지 추가)
@login_required
def create_meet_and_matches(request):
    if request.method == "POST" and is_manager(request.user):
        title = request.POST.get('title')
        meet_date = request.POST.get('date')
        selected_profile_ids = request.POST.getlist('profiles') 
        selected_courts = request.POST.getlist('courts') 

        if not selected_courts:
            messages.error(request, "최소 하나의 코트를 선택해야 합니다.")
            return redirect('dashboard')

        available_courts = [int(c) for c in selected_courts]

        if title and meet_date:
            # 🌟 신규 방어 로직: 이미 해당 날짜에 모임이 있는지 DB에서 확인
            if MonthlyMeet.objects.filter(date=meet_date).exists():
                messages.error(request, f"🚫 {meet_date} 에는 이미 생성된 모임이 있습니다. 다른 날짜를 선택해 주세요.")
                return redirect('dashboard')

            meet = MonthlyMeet.objects.create(title=title, date=meet_date)

            # 1. 기존 회원 리스트 가져오기
            players = list(Profile.objects.filter(id__in=selected_profile_ids))

            # 2. 게스트 데이터 처리 및 임시 프로필 생성
            guest_names = request.POST.getlist('guest_names[]')
            guest_groups = request.POST.getlist('guest_groups[]')
            guest_genders = request.POST.getlist('guest_genders[]')

            for n, g, gd in zip(guest_names, guest_groups, guest_genders):
                if n.strip(): 
                    guest_p = Profile.objects.create(
                        name=n.strip(), group=g, gender=gd, is_guest=True
                    )
                    players.append(guest_p)

            # 총 참가 인원이 2명 미만일 경우 방어 로직
            if len(players) < 2:
                meet.delete() # 생성했던 모임 삭제 롤백
                messages.error(request, "경기를 생성하려면 최소 2명 이상의 참가 인원(게스트 포함)이 필요합니다.")
                return redirect('dashboard')

            # 🌟 화면에서 입력받은 최소/최대 경기 수 적용
            try:
                MIN_MATCHES = int(request.POST.get('min_matches', 3))
                MAX_MATCHES = int(request.POST.get('max_matches', 4))
            except ValueError:
                MIN_MATCHES = 3
                MAX_MATCHES = 4
                
            if MIN_MATCHES > MAX_MATCHES:
                meet.delete() # 생성했던 모임 롤백
                messages.error(request, "최소 보장 경기 수가 최대 허용 경기 수보다 클 수 없습니다.")
                return redirect('dashboard')

            # 🌟 3. 다수 경기 스케줄링 (Round-Robin 기반) 및 제약 조건 적용 알고리즘
            MAX_ROUNDS = 100  # 무한 루프 방지용 (안전 장치 넉넉히 상향)

            match_counts = {p: 0 for p in players}      # 인원별 현재 매칭된 경기 수
            rest_counts = {p: 0 for p in players}       # 인원별 총 휴식(결장) 횟수
            consecutive_rest = {p: 0 for p in players}  # 연속으로 쉬고 있는 타임 수 (2연속 방지용)
            played_pairs = set()                        # 중복 매칭 방지용 Set (frozenset 이용)
            last_played_buffer = set()                  # 직전 타임(라운드)에 코트에 들어간 사람 기억
            
            scheduled_matches = []
            court_toggle = 2 # 2코트부터 번갈아가며 배정 시작

            for round_num in range(MAX_ROUNDS):
                # 🌟 종료 조건: 모든 인원이 허용된 최대 경기 수(MAX_MATCHES)에 도달했으면 즉시 스케줄링 종료
                if all(count >= MAX_MATCHES for count in match_counts.values()):
                    break
                
                # 이번 라운드(타임)에 코트에 배정할 후보들 (동점자 발생 시 편향 방지를 위해 먼저 섞음)
                available_players = list(players)
                random.shuffle(available_players)
                
                # 🌟 [알고리즘의 핵심] 배정 우선순위 정렬 (Sorting)
                # 리스트의 앞쪽에 있을수록 이번 타임에 코트에 들어갈 확률이 높아집니다.
                available_players.sort(key=lambda p: (
                    -consecutive_rest[p],                # 1순위: 연속 결장 횟수가 높은 사람 무조건 최우선 배치 (2연속 결장 강력 방어)
                    match_counts[p],                     # 2순위: 현재까지 경기 수가 '가장 적은' 사람 우선
                    1 if p in last_played_buffer else 0, # 3순위: 방금 전 타임에 뛴 사람은 후순위로 미룸 (최소 1경기 휴식 보장)
                    -rest_counts[p]                      # 4순위: 누적 휴식 횟수가 '많은' 사람을 우선 배치하여 불균형 해소
                ))
                
                round_matches = []
                players_used_this_round = set()
                
                # 한 라운드당 최대 2경기(2코트, 3코트) -> 4명 배정 시도
                for i in range(len(available_players)):
                    p1 = available_players[i]
                    if p1 in players_used_this_round: continue
                    if match_counts[p1] >= MAX_MATCHES: continue
                    
                    # p1과 매칭할 최적의 상대방(p2) 탐색
                    for j in range(i + 1, len(available_players)):
                        p2 = available_players[j]
                        if p2 in players_used_this_round: continue
                        if match_counts[p2] >= MAX_MATCHES: continue
                        
                        # 🌟 조건 1. 중복 매칭 방지 (No Rematch)
                        pair = frozenset([p1, p2])
                        if pair not in played_pairs:
                            round_matches.append((p1, p2))
                            players_used_this_round.update([p1, p2])
                            played_pairs.add(pair)
                            break # p1의 짝을 찾았으므로 다음 p1 탐색으로 넘어감
                            
                    # 2코트와 3코트가 모두 차면(2경기 4명) 이번 라운드 매칭 루프 탈출
                    if len(round_matches) == 2:
                        break
                        
                # 더 이상 매칭할 수 있는 짝이 없다면 전체 스케줄링 종료
                if not round_matches:
                    break
                    
                # 🌟 매치 객체 생성 및 코트 토글 분배
                for p1, p2 in round_matches:
                    handicap = calculate_handicap_logic(p1, p2)
                    scheduled_matches.append(Match(
                        meet=meet, player1=p1, player2=p2,
                        applied_handicap=handicap, court=court_toggle
                    ))
                    # 🌟 조건 2. 코트 토글 (2 -> 3 -> 2 -> 3)
                    court_toggle = 3 if court_toggle == 2 else 2
                    
                    # 경기 수 및 연속 휴식 카운트 초기화
                    match_counts[p1] += 1
                    match_counts[p2] += 1
                    consecutive_rest[p1] = 0
                    consecutive_rest[p2] = 0
                    
                # 🌟 조건 4. 휴식자(Bye) 카운트 업데이트
                resting_players = set(players) - players_used_this_round
                for rp in resting_players:
                    rest_counts[rp] += 1
                    consecutive_rest[rp] += 1
                    
                # 🌟 조건 3. 연속 출전 방지용 버퍼 업데이트
                last_played_buffer = players_used_this_round

            # DB에 일괄 저장 (Bulk Create로 퍼포먼스 향상)
            Match.objects.bulk_create(scheduled_matches)
            
            messages.success(request, f"'{title}' 모임과 대진표가 성공적으로 생성되었습니다! (스마트 스케줄링 적용)")
            
    return redirect('dashboard')

# 5. 점수 입력 (일반유저 1회 제한, 총관리자/사장님 무제한 수정 권한 적용)
@login_required
def record_score(request, match_id):
    match = get_object_or_404(Match, id=match_id)
    
    # 🌟 수정: 일반 유저가 완료된 경기 수정을 시도할 경우 차단 및 안내
    if match.is_completed and not is_manager(request.user):
        messages.error(request, "점수 수정은 사장님께 문의하세요.")
        return redirect('dashboard')
        
    if request.method == 'POST':
        match.p1_score = request.POST.get('p1_score')
        match.p2_score = request.POST.get('p2_score')
        match.is_completed = True
        
        # 🌟 추가: 최초 점수 입력자 저장 (이미 완료된 경기를 사장님이 수정할 땐 최초 입력자 유지)
        if not match.recorded_by:
            match.recorded_by = request.user
            
        match.save()
        messages.success(request, "점수가 성공적으로 기록되었습니다.")
        return redirect('dashboard')
    
# 6. 회원가입 함수
def signup(request):
    # 이미 로그인한 사람이 회원가입 페이지로 오면 메인으로 돌려보냄
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        # 🌟 [보안 수정] 공백 제거 및 글자 수 제한으로 DB 에러 방어
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        name = request.POST.get('name', '').strip()[:20] 
        group = request.POST.get('group')
        gender = request.POST.get('gender')

        # 🌟 필수 항목 누락 시 방어 로직
        if not username or not password or not name:
            messages.error(request, '필수 항목을 모두 입력해주세요.')
            return redirect('signup')

        # 아이디 중복 체크 방어 로직
        if User.objects.filter(username=username).exists():
            messages.error(request, '이미 사용 중인 아이디입니다. 다른 아이디를 입력해주세요.')
            return redirect('signup')

        # 1. Django 기본 User 계정 생성
        user = User.objects.create_user(username=username, password=password)
        
        # 2. 동호회 프로필(Profile) 생성 및 연결 (일반 가입자는 is_owner가 디폴트인 False로 들어감)
        Profile.objects.create(user=user, name=name, group=group, gender=gender)

        # 3. 가입 완료와 동시에 자동 로그인 처리 후 대시보드로 이동
        login(request, user)
        return redirect('dashboard')

    return render(request, 'matches/signup.html')

# 7. 회원 즉시 추가 기능 (총관리자/사장님 권한 적용)
@login_required
def add_member_by_admin(request):
    if request.method == 'POST' and is_manager(request.user):
        username = request.POST.get('username')
        password = request.POST.get('password')
        name = request.POST.get('name')
        group = request.POST.get('group')
        gender = request.POST.get('gender')

        if User.objects.filter(username=username).exists():
            messages.error(request, f"'{username}'(은)는 이미 사용 중인 아이디입니다.")
        else:
            # 1. 계정 생성
            user = User.objects.create_user(username=username, password=password)
            # 2. 프로필 생성
            Profile.objects.create(user=user, name=name, group=group, gender=gender)
            messages.success(request, f"🎉 {name} 회원({group}조)이 성공적으로 추가되었습니다!")
            
    return redirect('dashboard')

# (신규) 9. 오늘 모임 최종 마감 함수
@login_required
def finalize_meet(request, meet_id):
    # 🌟 [보안 수정] 권한 체크뿐만 아니라 GET 요청을 통한 CSRF 공격을 방어하기 위해 POST 요청만 허용
    if request.method != 'POST' or not is_manager(request.user):
        return redirect('dashboard')
    
    meet = get_object_or_404(MonthlyMeet, id=meet_id)
    meet.is_finalized = True
    meet.save()
    
    # 🌟 사장님께 마감 완료 알림을 띄워줍니다.
    messages.success(request, f"'{meet.title}' 모임이 마감되어 기록으로 보관되었습니다.")
    return redirect('dashboard')

# 10. 엑셀(CSV) 템플릿 다운로드 함수
@login_required
def download_member_template(request):
    if is_manager(request.user):
        # 엑셀에서 한글이 깨지지 않도록 utf-8-sig 포맷 사용
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="member_upload_template.csv"'
        
        writer = csv.writer(response)
        # 1행: 헤더 (안내문)
        writer.writerow(['접속아이디', '초기비밀번호', '실명', '조(A/B/C)', '성별(M/F)'])
        # 2행: 예시 데이터
        writer.writerow(['user01', '1234', '홍길동', 'B', 'M'])
        writer.writerow(['user02', '1234', '김유신', 'A', 'M'])
        writer.writerow(['user03', '1234', '유관순', 'C', 'F'])
        
        return response
    return redirect('dashboard')

# 11. 엑셀(CSV) 일괄 업로드 처리 함수
@login_required
def upload_members_bulk(request):
    if request.method == 'POST' and is_manager(request.user):
        csv_file = request.FILES.get('excel_file')
        
        if not csv_file or not csv_file.name.endswith('.csv'):
            messages.error(request, "CSV 파일을 업로드해주세요.")
            return redirect('dashboard')

        try:
            # 파일 읽기 및 파싱
            decoded_file = csv_file.read().decode('utf-8-sig').splitlines()
            reader = csv.reader(decoded_file)
            next(reader) # 첫 번째 줄(헤더)은 건너뜀
            
            success_count = 0
            for row in reader:
                if len(row) >= 5:
                    # 엑셀 데이터 추출 후 앞뒤 공백 제거
                    username = row[0].strip()
                    password = row[1].strip()
                    name = row[2].strip()
                    group = row[3].strip().upper()
                    gender = row[4].strip().upper()
                    
                    # 아이디가 존재하지 않을 때만 생성 (중복 방지)
                    if username and not User.objects.filter(username=username).exists():
                        user = User.objects.create_user(username=username, password=password)
                        Profile.objects.create(user=user, name=name, group=group, gender=gender)
                        success_count += 1
                        
            messages.success(request, f"🎉 총 {success_count}명의 회원이 성공적으로 일괄 추가되었습니다!")
        except Exception as e:
            messages.error(request, f"파일 처리 중 오류가 발생했습니다. 양식을 확인해주세요. (에러: {e})")
            
    return redirect('dashboard')

# 12. 경기 개별 수정 (선수, 코트 변경 등 - 사장님 전용)
@login_required
def update_match_detail(request, match_id):
    if request.method == "POST" and is_manager(request.user):
        match = get_object_or_404(Match, id=match_id)
        
        p1_id = request.POST.get('player1')
        p2_id = request.POST.get('player2')
        court = request.POST.get('court')
        
        # 선수 변경 반영
        match.player1 = get_object_or_404(Profile, id=p1_id)
        match.player2 = get_object_or_404(Profile, id=p2_id)
        match.court = int(court)
        
        # 점수가 함께 넘어온 경우 (수정 모드)
        p1_score = request.POST.get('p1_score')
        p2_score = request.POST.get('p2_score')
        if p1_score is not None and p2_score is not None:
            match.p1_score = int(p1_score)
            match.p2_score = int(p2_score)
            match.is_completed = True
            
        # 핸디캡 재계산 (선수가 바뀌었을 수 있으므로)
        match.applied_handicap = calculate_handicap_logic(match.player1, match.player2)
        match.save()
        
        messages.success(request, "경기 정보가 수정되었습니다.")
    return redirect('dashboard')

# 13. 독립된 전용 로그인 화면 뷰 (기존 custom_login 덮어쓰기)
def custom_login(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        user_id = request.POST.get('username')
        user_pw = request.POST.get('password')
        
        user = authenticate(request, username=user_id, password=user_pw)
        
        if user is not None:
            login(request, user)
            if 'login_failed_username' in request.session:
                del request.session['login_failed_username']
            return redirect('dashboard')
        else:
            request.session['login_failed_username'] = user_id
            messages.error(request, "아이디 또는 패스워드가 일치하지 않습니다.", extra_tags="login_error")
            return redirect('login') 
            
    login_failed_username = request.session.pop('login_failed_username', '')
    return render(request, 'matches/login.html', {'login_failed_username': login_failed_username})


# 14. 커스텀 로그아웃 (신규 추가)
def custom_logout(request):
    logout(request)
    return redirect('login')

# 17. 비밀번호 변경 기능
@login_required
def change_password(request):
    if request.method == 'POST':
        old_pw = request.POST.get('old_password', '')
        new_pw = request.POST.get('new_password', '').strip()
        user = request.user
        
        if not new_pw:
            messages.error(request, "새 비밀번호를 올바르게 입력해주세요.")
            return redirect('dashboard')
            
        if user.check_password(old_pw):
            user.set_password(new_pw)
            user.save()
            # 🌟 비밀번호가 바뀌어도 현재 세션을 유지(자동 로그아웃 방지)
            update_session_auth_hash(request, user)
            messages.success(request, "비밀번호가 성공적으로 변경되었습니다.")
        else:
            messages.error(request, "현재 비밀번호가 일치하지 않습니다.")
            
    return redirect('dashboard')

# 18. 경기 일정 엑셀 다운로드 (세로형 압축 버전 - 한 줄에 두 코트 나란히) - 사장님 전용
@login_required
def export_schedule_vertical(request, meet_id):
    if not is_manager(request.user):
        return redirect('dashboard')

    meet = get_object_or_404(MonthlyMeet, id=meet_id)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{meet.date} 일정표"

    # 스타일 설정
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True, size=13)
    center_align = Alignment(horizontal='center', vertical='center')

    # 1. 헤더 (1행)
    headers = ['순서', '2코트 경기', '3코트 경기']
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align
        # 헤더 배경색 (선택사항, 보기 좋게 연한 회색)
        from openpyxl.styles import PatternFill
        cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

    # 2. 데이터 가져오기
    matches_2 = list(Match.objects.filter(meet=meet, court=2).order_by('id'))
    matches_3 = list(Match.objects.filter(meet=meet, court=3).order_by('id'))
    max_len = max(len(matches_2), len(matches_3), 15) # 최소 15칸은 만들어둠
    tier_weight = {'A': 3, 'B': 2, 'C': 1}

    # 텍스트 가공 헬퍼 함수 (게스트, 핸디캡, 점수칸 합치기)
    def format_match_string(m):
        if not m: return ""
        p1_name = f"{m.player1.name}(g)" if m.player1.is_guest else m.player1.name
        p2_name = f"{m.player2.name}(g)" if m.player2.is_guest else m.player2.name
        h = m.applied_handicap
        if h > 0:
            if tier_weight[m.player1.group] < tier_weight[m.player2.group]:
                p1_name += f" (+{h})"
            elif tier_weight[m.player1.group] > tier_weight[m.player2.group]:
                p2_name += f" (+{h})"
            else: 
                if m.player1.gender == 'F' and m.player2.gender == 'M': p1_name += f" (+{h})"
                else: p2_name += f" (+{h})"
        # 이름들 뒤에 (   :   ) 점수 기입란 추가
        return f"{p1_name} vs {p2_name} (    :    )"

    # 3. 셀에 데이터 채우기
    for i in range(max_len):
        row_num = i + 2
        
        # 순서 열
        cell_no = ws.cell(row=row_num, column=1, value=i+1)
        cell_no.border = thin_border
        cell_no.alignment = center_align
        
        # 2코트 열
        m2 = matches_2[i] if i < len(matches_2) else None
        cell_m2 = ws.cell(row=row_num, column=2, value=format_match_string(m2))
        cell_m2.border = thin_border
        cell_m2.alignment = center_align

        # 3코트 열
        m3 = matches_3[i] if i < len(matches_3) else None
        cell_m3 = ws.cell(row=row_num, column=3, value=format_match_string(m3))
        cell_m3.border = thin_border
        cell_m3.alignment = center_align

        # 행 높이 조절 (펜으로 글씨 쓰기 편하게 살짝 넓게)
        ws.row_dimensions[row_num].height = 25

    # 4. 열 너비 조절
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 38  # 선수 이름 + 핸디캡 + 점수칸이 다 들어가게 넉넉히
    ws.column_dimensions['C'].width = 38

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="vertical_schedule_{meet.date}.xlsx"'
    wb.save(response)
    return response

# 19. 경기 일정 엑셀 다운로드 (양면 가로형 - 3코트 / 2코트) - 사장님 전용
@login_required
def export_schedule_horizontal(request, meet_id):
    if not is_manager(request.user):
        return redirect('dashboard')

    meet = get_object_or_404(MonthlyMeet, id=meet_id)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{meet.date} 경기 일정"

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A1:E1')
    ws['A1'] = '3 코트'
    ws.merge_cells('F1:J1')
    ws['F1'] = '2 코트'
    
    for cell in [ws['A1'], ws['F1']]:
        cell.font = Font(bold=True, size=14)
        cell.alignment = center_align

    headers = ['순서', '이름', 'vs', '이름', '비고']
    for col, val in enumerate(headers, 1): 
        cell = ws.cell(row=2, column=col, value=val)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    for col, val in enumerate(headers, 6): 
        cell = ws.cell(row=2, column=col, value=val)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    matches_3 = list(Match.objects.filter(meet=meet, court=3).order_by('id'))
    matches_2 = list(Match.objects.filter(meet=meet, court=2).order_by('id'))
    max_len = max(len(matches_3), len(matches_2), 15)
    
    tier_weight = {'A': 3, 'B': 2, 'C': 1}

    for i in range(max_len):
        row_num = i + 3
        
        # --- [좌측: 3 코트] ---
        ws.cell(row=row_num, column=1, value=i+1).border = thin_border
        if i < len(matches_3):
            m = matches_3[i]
            # 🌟 신규: 엑셀에도 게스트 (g) 표시 추가
            p1_name = f"{m.player1.name}(g)" if m.player1.is_guest else m.player1.name
            p2_name = f"{m.player2.name}(g)" if m.player2.is_guest else m.player2.name
            h = m.applied_handicap
            if h > 0:
                if tier_weight[m.player1.group] < tier_weight[m.player2.group]:
                    p1_name += f" (+{h})"
                elif tier_weight[m.player1.group] > tier_weight[m.player2.group]:
                    p2_name += f" (+{h})"
                else: 
                    if m.player1.gender == 'F' and m.player2.gender == 'M': p1_name += f" (+{h})"
                    else: p2_name += f" (+{h})"
            
            ws.cell(row=row_num, column=2, value=p1_name).border = thin_border
            ws.cell(row=row_num, column=3, value='vs').border = thin_border
            ws.cell(row=row_num, column=4, value=p2_name).border = thin_border
        else:
            for c in range(2, 5): ws.cell(row=row_num, column=c).border = thin_border
        ws.cell(row=row_num, column=5).border = thin_border 

        # --- [우측: 2 코트] ---
        ws.cell(row=row_num, column=6, value=i+1).border = thin_border
        if i < len(matches_2):
            m = matches_2[i]
            # 🌟 신규: 엑셀에도 게스트 (g) 표시 추가
            p1_name = f"{m.player1.name}(g)" if m.player1.is_guest else m.player1.name
            p2_name = f"{m.player2.name}(g)" if m.player2.is_guest else m.player2.name
            h = m.applied_handicap
            if h > 0:
                if tier_weight[m.player1.group] < tier_weight[m.player2.group]:
                    p1_name += f" (+{h})"
                elif tier_weight[m.player1.group] > tier_weight[m.player2.group]:
                    p2_name += f" (+{h})"
                else: 
                    if m.player1.gender == 'F' and m.player2.gender == 'M': p1_name += f" (+{h})"
                    else: p2_name += f" (+{h})"

            ws.cell(row=row_num, column=7, value=p1_name).border = thin_border
            ws.cell(row=row_num, column=8, value='vs').border = thin_border
            ws.cell(row=row_num, column=9, value=p2_name).border = thin_border
        else:
            for c in range(7, 10): ws.cell(row=row_num, column=c).border = thin_border
        ws.cell(row=row_num, column=10).border = thin_border 

        for c in range(1, 11):
            ws.cell(row=row_num, column=c).alignment = center_align

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="schedule_horizontal_{meet.date}.xlsx"'
    wb.save(response)
    return response

# 20. 인원 관리 페이지 전용 데이터 계산
@login_required
def member_management(request):
    if not is_manager(request.user):
        return redirect('dashboard')
    
    profiles = Profile.objects.filter(is_owner=False)
    regular_members = []
    guest_members = []
    
    for p in profiles:
        # 참여한 고유 모임(MonthlyMeet) 개수 계산
        participation_count = Match.objects.filter(
            Q(player1=p) | Q(player2=p), 
            meet__is_finalized=True
        ).values('meet').distinct().count()
        
        data = {
            'profile': p,
            'attendance_count': participation_count,
        }
        if p.is_guest:
            guest_members.append(data)
        else:
            regular_members.append(data)
    
    # 🌟 추가된 핵심 로직: 전체 명단을 참여 횟수(attendance_count)가 많은 순으로 정렬합니다.
    regular_members_sorted = sorted(regular_members, key=lambda x: x['attendance_count'], reverse=True)
    guest_members_sorted = sorted(guest_members, key=lambda x: x['attendance_count'], reverse=True)
    all_members_sorted = sorted(regular_members + guest_members, key=lambda x: x['attendance_count'], reverse=True)
    
    return render(request, 'matches/member_management.html', {
        'member_data': all_members_sorted,          # 전체 리스트
        'regular_data': regular_members_sorted,     # 동호회 정규 인원
        'guest_data': guest_members_sorted,         # 게스트 인원
        'attendance_rank': regular_members_sorted[:5],  # 상위 5명은 정규 인원에서만 선발
    })

# 21. 등급(조) 수정 처리 (기존과 동일)
@login_required
def update_member_rank(request, profile_id):
    if request.method == "POST" and is_manager(request.user):
        profile = get_object_or_404(Profile, id=profile_id)
        new_group = request.POST.get('group')
        if new_group in ['A', 'B', 'C']:
            profile.group = new_group
            profile.save()
            messages.success(request, f"{profile.name} 님의 등급이 {new_group}조로 변경되었습니다.")
    return redirect('dashboard')

# 22. 월별 경기 결과 엑셀 다운로드 (세로형 압축 결과 보고서) - 사장님 전용
@login_required
def export_meet_results(request, meet_id):
    if not is_manager(request.user):
        return redirect('dashboard')

    meet = get_object_or_404(MonthlyMeet, id=meet_id)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{meet.date} 경기결과"

    # 스타일 설정
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True, size=13)
    center_align = Alignment(horizontal='center', vertical='center')

    # 1. 헤더 (1행)
    headers = ['순서', '2코트 경기 결과', '3코트 경기 결과']
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align
        from openpyxl.styles import PatternFill
        cell.fill = PatternFill(start_color='EBF1DE', end_color='EBF1DE', fill_type='solid') # 결과용 연녹색 배경

    # 2. 데이터 가져오기
    matches_2 = list(Match.objects.filter(meet=meet, court=2).order_by('id'))
    matches_3 = list(Match.objects.filter(meet=meet, court=3).order_by('id'))
    max_len = max(len(matches_2), len(matches_3), 10) 
    tier_weight = {'A': 3, 'B': 2, 'C': 1}

    # 결과 텍스트 가공 헬퍼 함수 (기존 세로형 양식과 동일 + 점수 포함)
    def format_result_string(m):
        if not m: return ""
        p1_name = f"{m.player1.name}(g)" if m.player1.is_guest else m.player1.name
        p2_name = f"{m.player2.name}(g)" if m.player2.is_guest else m.player2.name
        h = m.applied_handicap
        if h > 0:
            if tier_weight[m.player1.group] < tier_weight[m.player2.group]:
                p1_name += f" (+{h})"
            elif tier_weight[m.player1.group] > tier_weight[m.player2.group]:
                p2_name += f" (+{h})"
            else: 
                if m.player1.gender == 'F' and m.player2.gender == 'M': p1_name += f" (+{h})"
                else: p2_name += f" (+{h})"
        
        # 점수 표시 (완료된 경우 점수 삽입, 미완료 시 빈 칸)
        s1 = m.p1_score if m.is_completed else "  "
        s2 = m.p2_score if m.is_completed else "  "
        return f"{p1_name}  {s1} : {s2}  {p2_name}"

    # 3. 셀에 데이터 채우기
    for i in range(max_len):
        row_num = i + 2
        
        # 순서 열
        cell_no = ws.cell(row=row_num, column=1, value=i+1)
        cell_no.border = thin_border
        cell_no.alignment = center_align
        
        # 2코트 결과
        m2 = matches_2[i] if i < len(matches_2) else None
        cell_m2 = ws.cell(row=row_num, column=2, value=format_result_string(m2))
        cell_m2.border = thin_border
        cell_m2.alignment = center_align

        # 3코트 결과
        m3 = matches_3[i] if i < len(matches_3) else None
        cell_m3 = ws.cell(row=row_num, column=3, value=format_result_string(m3))
        cell_m3.border = thin_border
        cell_m3.alignment = center_align

        ws.row_dimensions[row_num].height = 25

    # 4. 열 너비 조절 (기존과 동일하게 넉넉히)
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 45

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="match_results_{meet.date}.xlsx"'
    wb.save(response)
    return response

# 23. 🚨 변수 발생: 결장자 처리 및 스마트 재배치 (코트 균형 & 연속 휴식 보장)
@login_required
def handle_absentee_and_rebalance(request, meet_id):
    if not is_manager(request.user) or request.method != 'POST':
        return redirect('dashboard')
    
    absentee_id = request.POST.get('absentee_id')
    meet = get_object_or_404(MonthlyMeet, id=meet_id)
    
    # 1. 현재 미완료된 모든 경기를 가져옵니다.
    incomplete_matches = Match.objects.filter(meet=meet, is_completed=False)
    
    # 결장자를 제외한 정상적인 대진과, 결장자 때문에 상대방을 잃은 사람(orphans)을 분류합니다.
    valid_pairs = []
    orphans = []
    
    for m in incomplete_matches:
        p1_id = str(m.player1.id)
        p2_id = str(m.player2.id)
        
        if p1_id == absentee_id:
            orphans.append(m.player2)
        elif p2_id == absentee_id:
            orphans.append(m.player1)
        else:
            valid_pairs.append((m.player1, m.player2))
            
    # 기존 미완료 경기들은 삭제합니다.
    incomplete_matches.delete()
    
    # 2. 상대방을 잃은 사람들끼리 랜덤으로 새로운 짝을 맺어줍니다.
    random.shuffle(orphans)
    while len(orphans) >= 2:
        valid_pairs.append((orphans.pop(), orphans.pop()))
        
    # (홀수가 남아 1명이 남는다면, 부전승 처리 개념으로 이번 라운드는 쉬게 됩니다)

    # 3. 2코트/3코트 1:1 균형 및 연속 출전 방지 정렬 알고리즘
    scheduled_matches = []
    last_played_buffer = [] # 방금 전 코트에 들어간 4명(2코트 2명, 3코트 2명)을 기억하여 연속 출전 방지
    court_toggle = 2 # 2코트부터 배정 시작
    
    while valid_pairs:
        best_idx = 0
        # 대기 중인 매치 중, 두 선수 모두 방금 전 경기를 뛰지 않은 최적의 매치를 찾습니다.
        for i, (p1, p2) in enumerate(valid_pairs):
            if p1 not in last_played_buffer and p2 not in last_played_buffer:
                best_idx = i
                break
                
        # 매치 확정 및 리스트에서 제거
        p1, p2 = valid_pairs.pop(best_idx)
        
        # 🌟 핵심 추가: 재배치되어 만난 두 사람의 핸디캡을 다시 계산합니다.
        handicap = calculate_handicap_logic(p1, p2)
        
        # 새로운 경기 객체 생성 (저장 대기)
        scheduled_matches.append(Match(
            meet=meet, 
            court=court_toggle, 
            player1=p1, 
            player2=p2,
            applied_handicap=handicap  # 👈 계산된 핸디캡을 DB에 저장하도록 추가!
        ))
        
        # 휴식 보장을 위해 방금 뛴 사람 명단 업데이트 (최대 4명 기억)
        last_played_buffer.extend([p1, p2])
        if len(last_played_buffer) > 4:
            last_played_buffer = last_played_buffer[-4:]
            
        # 🌟 2코트와 3코트를 정확히 번갈아가며 배정 (비율 균형 맞춤)
        court_toggle = 3 if court_toggle == 2 else 2
        
    # 데이터베이스에 일괄 저장
    Match.objects.bulk_create(scheduled_matches)
    
    messages.success(request, "🚨 결장자가 제외되고, 핸디캡과 코트 비율을 고려하여 대진이 완벽하게 재배치되었습니다!")
    return redirect('dashboard')

# 24. 📡 실시간 점수 데이터 통신 (AJAX 전용)
@login_required
def get_live_scores(request, meet_id):
    matches = Match.objects.filter(meet_id=meet_id)
    data = []
    for m in matches:
        data.append({
            'id': m.id,
            'p1_score': m.p1_score if m.p1_score is not None else '-',
            'p2_score': m.p2_score if m.p2_score is not None else '-',
            'is_completed': m.is_completed,
            'recorded_by': escape(m.recorded_by.profile.name) if m.recorded_by else '-' # 🌟 [보안 수정] 이름 XSS 방어
        })
    return JsonResponse({'matches': data})

# 25. 📅 모임 일정 취소 (DB에서 삭제)
@login_required
def cancel_meeting(request, meet_id):
    # 🌟 [보안 수정] 권한 체크뿐만 아니라 GET 요청을 통한 CSRF 공격을 방어하기 위해 POST 요청만 허용
    if request.method != 'POST' or not is_manager(request.user):
        return redirect('dashboard')
    
    meet = get_object_or_404(MonthlyMeet, id=meet_id)
    # CASCADE 설정으로 인해 연결된 Match들도 함께 삭제됩니다.
    meet.delete()
    
    messages.success(request, "⚠️ 모임 일정이 취소되고 모든 대진표가 삭제되었습니다.")
    return redirect('dashboard')

# 26. ⚙️ 핸디캡 수동 일괄 수정 (사장님 전용 오버라이드)
@login_required
def update_handicaps(request, meet_id):
    if not is_manager(request.user) or request.method != 'POST':
        return redirect('dashboard')
    
    meet = get_object_or_404(MonthlyMeet, id=meet_id)
    matches = Match.objects.filter(meet=meet)
    
    update_count = 0
    for m in matches:
        # 화면에서 넘어온 'handicap_경기번호' 값을 찾습니다.
        handicap_val = request.POST.get(f'handicap_{m.id}')
        
        if handicap_val is not None and handicap_val.isdigit():
            new_handicap = int(handicap_val)
            
            # 기존 점수와 다르다면 새 점수로 덮어씁니다.
            if m.applied_handicap != new_handicap:
                m.applied_handicap = new_handicap
                m.save()
                update_count += 1
                
    if update_count > 0:
        messages.success(request, f"✅ {update_count}건의 매치 핸디캡이 수동으로 변경되었습니다!")
        
    return redirect('dashboard')

# 27. 🚀 게스트를 동호회 정규 인원으로 승격 (사장님 전용)
@login_required
def promote_guest(request, profile_id):
    if not is_manager(request.user) or request.method != 'POST':
        return redirect('dashboard')
    
    profile = get_object_or_404(Profile, id=profile_id, is_guest=True)
    profile.is_guest = False
    profile.save()
    
    messages.success(request, f"🎉 {profile.name} 님이 동호회 정규 인원으로 승격되었습니다!")
    return redirect('member_management')

# 28. 과거 경기 기록 일괄 업로드 템플릿 다운로드
@login_required
def download_match_template(request):
    if is_manager(request.user):
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="match_upload_template.csv"'
        
        writer = csv.writer(response)
        # 1행: 헤더 (안내문)
        writer.writerow(['모임일자(YYYY-MM-DD)', '모임이름(예: 1월정기모임)', '코트(2/3)', '선수1(실명)', '선수2(실명)', '선수1점수', '선수2점수'])
        # 2행: 예시 데이터
        writer.writerow(['2024-01-15', '1월 정기 모임', '2', '홍길동', '김철수', '11', '8'])
        writer.writerow(['2024-01-15', '1월 정기 모임', '3', '이영희', '박민수', '9', '11'])
        
        return response
    return redirect('dashboard')

# 29. 과거 경기 기록 엑셀(CSV) 일괄 업로드 처리
@login_required
def upload_matches_bulk(request):
    if request.method == 'POST' and is_manager(request.user):
        csv_file = request.FILES.get('excel_file')
        
        if not csv_file or not csv_file.name.endswith('.csv'):
            messages.error(request, "CSV 파일을 업로드해주세요.")
            return redirect('dashboard')

        try:
            decoded_file = csv_file.read().decode('utf-8-sig').splitlines()
            reader = csv.reader(decoded_file)
            next(reader) # 첫 번째 줄(헤더) 건너뜀
            
            success_count = 0
            for row in reader:
                if len(row) >= 7:
                    date_str, title, court, p1_name, p2_name, p1_score, p2_score = [
                        col.replace('\ufeff', '').replace('“', '').replace('”', '').replace('"', '').strip() 
                        for col in row[:7]
                    ]
                    
                    # DB에서 선수 프로필 찾기 (동명이인 처리 로직이 필요하다면 고도화 가능)
                    p1 = Profile.objects.filter(name=p1_name).first()
                    p2 = Profile.objects.filter(name=p2_name).first()
                    
                    if not p1 or not p2:
                        continue # 등록되지 않은 이름이면 스킵
                        
                    # 모임(MonthlyMeet) 찾기 또는 새로 생성 (과거 기록이므로 무조건 마감 처리)
                    meet, created = MonthlyMeet.objects.get_or_create(
                        date=date_str,
                        defaults={'title': title, 'is_finalized': True}
                    )
                    
                    # 매치(Match) 객체 생성
                    handicap = calculate_handicap_logic(p1, p2)
                    Match.objects.create(meet=meet, court=int(court), player1=p1, player2=p2, p1_score=int(p1_score), p2_score=int(p2_score), applied_handicap=handicap, is_completed=True, recorded_by=request.user)
                    success_count += 1
                    
            messages.success(request, f"🎉 총 {success_count}건의 과거 경기 기록이 성공적으로 일괄 추가되었습니다!")
        except Exception as e:
            messages.error(request, f"파일 처리 중 오류가 발생했습니다. 양식을 확인해주세요. (에러: {e})")
            
    return redirect('dashboard')

# --- 공지사항 관련 뷰 ---

# 🌟 [신규] 공지사항 목록 (팝업용, AJAX)
@login_required
def notice_list(request):
    search_keyword = request.GET.get('keyword', '')
    # 🌟 [수정] 먼저 댓글 수만 조회하고, 순번은 파이썬에서 직접 매깁니다.
    notice_list_qs = Notice.objects.annotate(comment_count=Count('comments'))

    if search_keyword:
        notice_list_qs = notice_list_qs.filter(title__icontains=search_keyword)

    # 🌟 일반 공지사항의 총 개수만 계산합니다.
    total_normal_count = notice_list_qs.filter(is_important=False).count()
    notices_ordered = notice_list_qs.order_by('-is_important', '-created_at')
    
    # 직렬화
    data = []
    normal_index = 0
    for notice in notices_ordered:
        if notice.is_important:
            display_no = "공지"
        else:
            display_no = total_normal_count - normal_index
            normal_index += 1
            
        data.append({
            'id': notice.id,
            'display_id': display_no, # 🌟 중요 공지는 '공지', 일반 글은 빈틈없는 순차 번호
            'no': display_no, # 프론트엔드 호환성 유지
            'title': escape(notice.title),
            'author': escape(notice.get_author_name()),
            'created_at': notice.created_at.strftime('%Y-%m-%d'),
            'is_important': notice.is_important,
            'comment_count': notice.comment_count,
        })
    
    return JsonResponse({'notices': data})

# 🌟 [신규] 공지사항 상세 정보 (팝업용, AJAX)
@login_required
def notice_detail(request, notice_id):
    notice = get_object_or_404(Notice, id=notice_id)
    # 🌟 조회수 1 증가 (F 객체 사용으로 동시성 문제 방지)
    notice.view_count = F('view_count') + 1
    notice.save(update_fields=['view_count'])
    notice.refresh_from_db() # DB에 저장된 최신 값을 다시 불러옴

    comments = notice.comments.select_related('author__profile').order_by('created_at')

    # 직렬화
    notice_data = {
        'id': notice.id,
        'title': notice.title,
        'content': escape(notice.content), # 🌟 [보안 수정] XSS 공격 방지를 위해 HTML을 escape 처리합니다.
        'author': notice.get_author_name(),
        'created_at': notice.created_at.strftime('%Y-%m-%d %H:%M'),
        'is_important': notice.is_important,
        'location_name': escape(notice.location_name) if notice.location_name else '', # 🌟 [보안 수정] 부가 정보 XSS 방어
        'author_display_name': escape(notice.author_display_name) if notice.author_display_name else '',
        'can_edit': is_manager(request.user), # 수정/삭제 권한 여부
        'view_count': notice.view_count, # 🌟 조회수 추가
        'comment_count': comments.count(), # 🌟 댓글 수 추가
    }
    
    comments_data = []
    for comment in comments:
        comments_data.append({
            'id': comment.id,
            'author': comment.author.profile.name if hasattr(comment.author, 'profile') else comment.author.username,
            'content': escape(comment.content), # 🌟 [보안 수정] XSS 공격 방지를 위해 HTML을 escape 처리합니다.
            'created_at': comment.created_at.strftime('%Y-%m-%d %H:%M'),
            'can_delete': request.user == comment.author or is_manager(request.user)
        })

    return JsonResponse({'notice': notice_data, 'comments': comments_data})

# 🌟 [신규] 공지사항 생성/수정
@login_required
def notice_save(request):
    if not is_manager(request.user):
        messages.error(request, "권한이 없습니다.")
        return redirect('dashboard')

    if request.method == 'POST':
        notice_id = request.POST.get('notice_id')
        title = request.POST.get('title')
        content = request.POST.get('content')
        is_important = request.POST.get('is_important') == 'on'
        location_name = request.POST.get('location_name')
        author_display_name = request.POST.get('author_display_name')

        if notice_id: # 수정
            notice = get_object_or_404(Notice, id=notice_id)
            notice.title, notice.content, notice.is_important, notice.location_name = title, content, is_important, location_name
            if request.user.is_superuser:
                notice.author_display_name = author_display_name
            notice.save()
            messages.success(request, "공지사항이 수정되었습니다.")
        else: # 생성
            notice = Notice.objects.create(author=request.user, title=title, content=content, is_important=is_important, location_name=location_name)
            if request.user.is_superuser and author_display_name:
                notice.author_display_name = author_display_name
                notice.save()
            messages.success(request, "공지사항이 등록되었습니다.")
    return redirect('dashboard')

# 🌟 [신규] 공지사항 삭제
@login_required
def notice_delete(request, notice_id):
    if request.method == 'POST' and is_manager(request.user):
        get_object_or_404(Notice, id=notice_id).delete()
        messages.success(request, "공지사항이 삭제되었습니다.")
    return redirect('dashboard')

# 🌟 [신규] 댓글 추가/삭제
@login_required
def add_notice_comment(request, notice_id):
    if request.method == 'POST' and request.POST.get('comment_content'):
        NoticeComment.objects.create(notice_id=notice_id, author=request.user, content=request.POST.get('comment_content'))
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def delete_notice_comment(request, comment_id):
    if request.method == 'POST':
        comment = get_object_or_404(NoticeComment, id=comment_id)
        if request.user == comment.author or is_manager(request.user):
            comment.delete()
            return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)